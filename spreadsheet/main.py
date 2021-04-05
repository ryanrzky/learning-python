import openpyxl

invFile = openpyxl.load_workbook("inventory.xlsx")
productList = invFile["Sheet1"]

productPerSupplier = {}
totalValuePerSupplier = {}
productUnder10 = {}

for productRow in range(2, productList.max_row + 1):
    supplierName = productList.cell(productRow, 4).value
    inventory = productList.cell(productRow, 2).value
    price = productList.cell(productRow, 3).value
    productNum = productList.cell(productRow, 1).value
    inventoryPrice = productList.cell(productRow, 5)
    
    # calculation number of product per supplier
    if supplierName in productPerSupplier:
        currentNumberProduct = productPerSupplier.get(supplierName)
        productPerSupplier[supplierName] = currentNumberProduct + 1
    else:
        productPerSupplier[supplierName] = 1

    # calculation total valuer of inventory per supplier
    if supplierName in totalValuePerSupplier:
        currentTotalValue = totalValuePerSupplier.get(supplierName)
        totalValuePerSupplier[supplierName] = currentTotalValue + inventory * price
    else:
        totalValuePerSupplier[supplierName] = inventory * price

    # logic product with inventory less than 10
    if inventory < 10:
        productUnder10[int(productNum)] = int(inventory)

    # add value for total inventory price
    inventoryPrice.value = inventory * price
            


print(f"product per suppler {productPerSupplier}")
print(f"total valuer per supplier {totalValuePerSupplier}")
print(f"product under 10 {productUnder10}")

# save update file
invFile.save("inventory-update.xlsx")